from __future__ import annotations

import csv
from pathlib import Path

from settlement_reconciliation.errors import LoadError
from settlement_reconciliation.models import RawRow


def load_tabular_file(path: str | Path) -> list[RawRow]:
    source = Path(path)
    suffix = source.suffix.lower()
    if suffix == ".csv":
        return load_csv(source)
    if suffix in {".xlsx", ".xlsm"}:
        return load_xlsx(source)
    raise LoadError(f"unsupported file type: {source.suffix}")


def load_csv(path: str | Path) -> list[RawRow]:
    source = Path(path)
    try:
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            return [dict(row) for row in csv.DictReader(handle)]
    except OSError as exc:
        raise LoadError(f"failed to read CSV file: {source}") from exc


def load_xlsx(path: str | Path) -> list[RawRow]:
    source = Path(path)
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError as exc:
        raise LoadError("XLSX input requires installing the 'xlsx' extra") from exc

    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    except OSError as exc:
        raise LoadError(f"failed to read XLSX file: {source}") from exc

    if not rows:
        return []

    headers = ["" if value is None else str(value).strip() for value in rows[0]]
    output: list[RawRow] = []
    for row in rows[1:]:
        output.append(
            {
                header: "" if value is None else str(value).strip()
                for header, value in zip(headers, row, strict=False)
                if header
            }
        )
    return output

